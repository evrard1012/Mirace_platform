"""
Profilage 2 : clustering hiérarchique multivarié (Soft-DTW) feuille par feuille
==============================================================================
• Lecture d’un classeur Excel
• Calcul Soft-DTW entre pays (toutes les variables à partir de la 3ᵉ colonne)
• CAH (Ward) + dendrogramme + DataFrame des clusters
=============================================================================="""

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler
from tslearn.metrics import soft_dtw
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
import openpyxl                       # lecture Excel
from pathlib import Path
import tempfile

from tslearn.barycenters import softdtw_barycenter

# Paramètre global Soft-DTW
GAMMA = 1.0


# ──────────────────────────────────────────────────────────────────────────────
# 1. Distance Soft-DTW multivariée
# ──────────────────────────────────────────────────────────────────────────────
def _distance_matrix(df: pd.DataFrame) -> tuple[np.ndarray, list[str]]:
    """Retourne (matrice NxN, liste_pays) en normalisant chaque variable."""
    pays = df.iloc[:, 0].unique()
    feat_df = df.iloc[:, 2:]                   # variables à partir de la 3ᵉ colonne

    # Dict Pays → matrice (Années × Variables)
    scaler = StandardScaler()
    tensor = []
    for p in pays:
        mat = scaler.fit_transform(feat_df[df.iloc[:, 0] == p].values)
        tensor.append(mat)

    n = len(pays)
    dist = np.zeros((n, n))
    for i in range(n):
        for j in range(i + 1, n):
            d = soft_dtw(tensor[i], tensor[j], gamma=GAMMA)
            dist[i, j] = dist[j, i] = d
    return dist, list(pays)


# ──────────────────────────────────────────────────────────────────────────────
# 2. Dendrogramme + DataFrame de clusters
# ──────────────────────────────────────────────────────────────────────────────
def _cluster_and_plot(dist: np.ndarray,
                      pays: list[str],
                      n_groups: int,
                      title: str):
    """Retourne (fig, df_clusters trié)."""
    Z = linkage(dist, method="ward")

    # Figure en mémoire (utile pour Streamlit)
    fig, ax = plt.subplots(figsize=(12, 5))
    dendrogram(Z, labels=pays, leaf_rotation=90, ax=ax)
    ax.set_title(f"Dendrogramme (Soft-DTW) – {title}")
    ax.set_xlabel("Pays")
    ax.set_ylabel("Distance")
    plt.tight_layout()

    clusters = fcluster(Z, n_groups, criterion="maxclust")
    df_clust = (
        pd.DataFrame({"Pays": pays, "Cluster": clusters})
        .sort_values("Cluster")
        .reset_index(drop=True)
    )
    return fig, df_clust


# ──────────────────────────────────────────────────────────────────────────────
# 3. Fonction publique : analyse de toutes les feuilles Excel
# ──────────────────────────────────────────────────────────────────────────────
def analyse_profilage2_2(file_path: str | Path, n_groups: int = 4) -> dict:
    """
    Analyse chaque feuille Excel :
    retourne { nom_feuille → (fig, df_clusters) }
    """
    results = {}
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet in wb.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet, header=0, engine="openpyxl")
        dist, pays = _distance_matrix(df)
        fig, df_clusters = _cluster_and_plot(dist, pays, n_groups, title=sheet)
        results[sheet] = (fig, df_clusters)
    return results

def analyse_profilage2(file_or_upload: str | Path, n_groups: int = 4) -> dict:
    """
    Analyse chaque feuille Excel :
    - Si file_or_upload est un UploadedFile (streamlit), il est d'abord copié temporairement
    - Sinon, file_or_upload est un chemin local
    Retourne : { nom_feuille → (fig, df_clusters) }
    """
    results = {}

    # Cas fichier uploadé
    if hasattr(file_or_upload, "read"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_or_upload.read())
            tmp_path = tmp.name
    else:
        tmp_path = str(file_or_upload)  # Compatible avec str ou Path

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    for sheet in wb.sheetnames:
        df = pd.read_excel(tmp_path, sheet_name=sheet, header=0, engine="openpyxl")
        dist, pays = _distance_matrix(df)
        fig, df_clusters = _cluster_and_plot(dist, pays, n_groups, title=sheet)
        results[sheet] = (fig, df_clusters)

    return results
    

def calcul_centroids(data: pd.DataFrame, df_clusters: pd.DataFrame) -> dict:
    """
    Calcule le barycentre Soft-DTW de chaque cluster pour une feuille Excel.
    ------------------------------------------------------------------------
    data        : DataFrame brut | colonnes = Pays, Année, Facteur1, Facteur2…
    df_clusters : DataFrame 'Pays' + 'Cluster' produit par Profilage 2
    Retour      : { cluster_id -> ndarray shape (n_years, n_factors) }
    """

    # 0️⃣  préparation : identifier variables et années
    factors = data.columns[2:]
    years   = np.sort(data.iloc[:, 1].unique())          # tri années
    n_years, n_factors = len(years), len(factors)

    # 1️⃣  construire un tensor « pays → (années × facteurs) »
    tensor   = {}                                        # Pays -> array (n_years, n_factors)
    scaler   = StandardScaler()
    for pays in data.iloc[:, 0].unique():
        sub = data[data.iloc[:, 0] == pays]
        sub = sub.sort_values(data.columns[1])           # tri par année
        values = sub[factors].values                     # shape (n_years, n_factors)
        tensor[pays] = scaler.fit_transform(values)      # normalisation

    # 2️⃣  créer un dict cluster -> liste(numpy array)
    clusters = (
        data.iloc[:, :1]                                  # Pays
        .merge(df_clusters, on="Pays", how="left")        # ajoute Cluster
        .groupby("Cluster")["Pays"]
        .apply(list)
        .to_dict()
    )

    # 3️⃣  barycentre Soft-DTW pour chaque cluster
    centroids = {}
    for cl_id, pays_list in clusters.items():
        # empilement des séries (nb_pays, n_years, n_factors)
        series = np.stack([tensor[p] for p in pays_list])

        # barycentre indépendant sur chaque facteur
        centroid = np.array([
            softdtw_barycenter(series[:, :, f], gamma=GAMMA)
            for f in range(n_factors)
        ]).T                                             # shape (n_years, n_factors)

        centroids[cl_id] = centroid

    return centroids, years, factors

def centroid_unique(df_concat: pd.DataFrame) -> tuple[np.ndarray, np.ndarray, list]:
    """
    Calcule le barycentre Soft-DTW (toutes lignes → 1 seul centroïde).
    -----------------------------------------------------------------
    df_concat doit contenir les colonnes : 'Pays', 'Années', <facteurs…>
    Retour : (centroid, years, factors)
             centroid  shape = (n_years, n_factors)
    """
    factors = df_concat.columns[2:]                 # >2 = variables
    years   = np.sort(df_concat["Années"].unique())
    n_years, n_fac = len(years), len(factors)
    df_concat[factors] = df_concat[factors].fillna(0)

    # 1) tensor 3-D (nb_pays, n_years, n_factors)
    scaler = StandardScaler()
    series_list = []
    for p in df_concat["Pays"].unique():
        sub = df_concat[df_concat["Pays"] == p].sort_values("Années")
        X   = scaler.fit_transform(sub[factors].values)   # normalise
        series_list.append(X)

    stack = np.stack(series_list)             # shape (nb_pays, n_years, n_factors)

    # 2) barycentre Soft-DTW variable par variable
    centroid = np.array([
        softdtw_barycenter(stack[:, :, f], gamma=GAMMA).ravel() for f in range(n_fac)
    ]).T                                        # (n_years, n_factors)
    

    return centroid, years, list(factors)